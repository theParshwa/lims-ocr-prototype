"""
excel_generator.py - Generate a multi-sheet LIMS Load Sheet Excel file.

Generates all 30 standard LIMS Load Sheet tabs with styling:
  - Header row: blue background, white bold text
  - Low-confidence cells: yellow highlight
  - Validation error cells: red highlight
  - Auto column widths
  - Freeze top row
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models.schemas import ExtractionResult

logger = logging.getLogger(__name__)

# ── Colour constants ──────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="1F5C9E")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_LOW_CONF_FILL = PatternFill("solid", fgColor="FFEB9C")   # yellow
_ERROR_FILL = PatternFill("solid", fgColor="FFC7CE")      # red
_BORDER_SIDE = Side(style="thin", color="CCCCCC")
_CELL_BORDER = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE,
    top=_BORDER_SIDE, bottom=_BORDER_SIDE,
)
_NORMAL_FONT = Font(name="Calibri", size=10)


# ── Sheet definitions — all 30 LIMS Load Sheet tabs ─────────────────────────

SHEET_SPECS: list[dict[str, Any]] = [
    {
        "name": "ANALYSIS_TYPES",
        "data_attr": "analysis_types",
        "columns": ["NAME", "DESCRIPTION", "GROUP_NAME"],
        "field_map": ["name", "description", "group_name"],
    },
    {
        "name": "COMMON_NAME",
        "data_attr": "common_names",
        "columns": ["NAME", "DESCRIPTION", "GROUP_NAME"],
        "field_map": ["name", "description", "group_name"],
    },
    {
        "name": "ANALYSIS",
        "data_attr": "analysis",
        "columns": ["NAME", "VERSION", "GROUP_NAME", "ACTIVE", "REPORTED_NAME", "COMMON_NAME", "ANALYSIS_TYPE", "DESCRIPTION"],
        "field_map": ["name", "version", "group_name", "active", "reported_name", "common_name", "analysis_type", "description"],
    },
    {
        "name": "COMPONENT",
        "data_attr": "components",
        "columns": ["ANALYSIS", "NAME", "NUM_REPLICATES", "VERSION", "ORDER_NUMBER", "RESULT_TYPE", "UNITS", "MINIMUM", "MAXIMUM"],
        "field_map": ["analysis", "name", "num_replicates", "version", "order_number", "result_type", "units", "minimum", "maximum"],
    },
    {
        "name": "UNITS",
        "data_attr": "units",
        "columns": ["UNIT_CODE", "DESCRIPTION", "DISPLAY_STRING", "GROUP_NAME"],
        "field_map": ["unit_code", "description", "display_string", "group_name"],
    },
    {
        "name": "PRODUCT",
        "data_attr": "products",
        "columns": [
            "PRODUCT", "VERSION", "SAMPLING_POINT", "GRADE", "ORDER_NUMBER",
            "DESCRIPTION", "CONTINUE_CHECKING", "TEST_LIST", "ALWAYS_CHECK",
            "T_PH_SAME_LOT", "T_PH_STATUS1", "T_PH_STATUS2", "T_PH_RECERT",
            "T_USP_SECONDARY", "TEST_LOCATION", "REQD_VOLUME", "ALIQUOT_GROUP",
            "C_SPEC_VARIATION", "C_TEST_GROUP",
        ],
        "field_map": [
            "product", "version", "sampling_point", "grade", "order_number",
            "description", "continue_checking", "test_list", "always_check",
            "t_ph_same_lot", "t_ph_status1", "t_ph_status2", "t_ph_recert",
            "t_usp_secondary", "test_location", "reqd_volume", "aliquot_group",
            "c_spec_variation", "c_test_group",
        ],
    },
    {
        "name": "T_PH_GRADE",
        "data_attr": "tph_grades",
        "columns": ["NAME", "DESCRIPTION", "GROUP_NAME", "ACTIVE", "CODE"],
        "field_map": ["name", "description", "group_name", "active", "code"],
    },
    {
        "name": "SAMPLING_POINT",
        "data_attr": "sampling_points",
        "columns": ["NAME", "DESCRIPTION", "GROUP_NAME"],
        "field_map": ["name", "description", "group_name"],
    },
    {
        "name": "PRODUCT_GRADE",
        "data_attr": "product_grades",
        "columns": ["PRODUCT", "VERSION", "GRADE", "ORDER_NUMBER", "DESCRIPTION", "CONTINUE_CHECKING", "TEST_LIST", "ALWAYS_CHECK"],
        "field_map": ["product", "version", "grade", "order_number", "description", "continue_checking", "test_list", "always_check"],
    },
    {
        "name": "PROD_GRADE_STAGE",
        "data_attr": "prod_grade_stages",
        "columns": [
            "PRODUCT", "VERSION", "SAMPLING_POINT", "GRADE", "STAGE", "ANALYSIS",
            "ORDER_NUMBER", "DESCRIPTION", "SPEC_TYPE", "NUM_REPS", "PARTIAL",
            "EXT_LINK", "REPORTED_NAME", "VARIATION", "REQUIRED",
        ],
        "field_map": [
            "product", "version", "sampling_point", "grade", "stage", "analysis",
            "order_number", "description", "spec_type", "num_reps", "partial",
            "ext_link", "reported_name", "variation", "required",
        ],
    },
    {
        "name": "PRODUCT_SPEC",
        "data_attr": "product_specs",
        "columns": [
            "PRODUCT", "VERSION", "SAMPLING_POINT", "GRADE", "STAGE", "ANALYSIS",
            "COMPONENT", "UNITS", "ROUND", "RULE_TYPE", "SPEC_RULE",
            "MIN_VALUE", "MAX_VALUE", "TEXT_VALUE", "CLASS", "REQUIRED",
        ],
        "field_map": [
            "product", "version", "sampling_point", "grade", "stage", "analysis",
            "component", "units", "round", "rule_type", "spec_rule",
            "min_value", "max_value", "text_value", "class_", "required",
        ],
    },
    {
        "name": "T_PH_ITEM_CODE",
        "data_attr": "tph_item_codes",
        "columns": [
            "T_PH_ITEM_CODE", "SUPPLIER", "ACTIVE", "STATUS1", "STATUS2",
            "ORDER_NUMBER", "RETEST_INTERVAL", "EXPIRY_INTERVAL",
            "FULL_TEST_FREQ", "LOTS_TO_GO", "DATE_LAST_TESTED",
        ],
        "field_map": [
            "t_ph_item_code", "supplier", "active", "status1", "status2",
            "order_number", "retest_interval", "expiry_interval",
            "full_test_freq", "lots_to_go", "date_last_tested",
        ],
    },
    {
        "name": "T_PH_ITEM_CODE_SPEC",
        "data_attr": "tph_item_code_specs",
        "columns": ["SPEC_CODE", "T_PH_ITEM_CODE", "PRODUCT_SPEC", "SPEC_CLASS", "ORDER_NUMBER", "GRADE", "COMMON_GRADE"],
        "field_map": ["spec_code", "t_ph_item_code", "product_spec", "spec_class", "order_number", "grade", "common_grade"],
    },
    {
        "name": "T_PH_ITEM_CODE_SUPP",
        "data_attr": "tph_item_code_supps",
        "columns": ["T_PH_ITEM_CODE", "SUPPLIER", "ACTIVE", "STATUS1", "STATUS2", "ORDER_NUMBER"],
        "field_map": ["t_ph_item_code", "supplier", "active", "status1", "status2", "order_number"],
    },
    {
        "name": "T_PH_SAMPLE_PLAN",
        "data_attr": "tph_sample_plans",
        "columns": ["NAME", "VERSION", "ACTIVE", "DESCRIPTION", "GROUP_NAME"],
        "field_map": ["name", "version", "active", "description", "group_name"],
    },
    {
        "name": "T_PH_SAMPLE_PLAN_EN",
        "data_attr": "tph_sample_plan_entries",
        "columns": [
            "T_PH_SAMPLE_PLAN", "ENTRY_CODE", "VERSION", "DESCRIPTION",
            "ORDER_NUMBER", "SPEC_TYPE", "SPEC_STATUS", "T_PH_SAMPLE_TYPE",
            "LOG_SAMPLE", "CREATE_INVENTORY", "RETAINED_SAMPLE", "STABILITY",
            "INITIAL_STATUS", "INDIC_SAMPLES", "QUANTITY", "UNITS",
            "C_REANALYSIS_QTY", "C_STM_QUANTITY",
        ],
        "field_map": [
            "t_ph_sample_plan", "entry_code", "version", "description",
            "order_number", "spec_type", "spec_status", "t_ph_sample_type",
            "log_sample", "create_inventory", "retained_sample", "stability",
            "initial_status", "indic_samples", "quantity", "units",
            "c_reanalysis_qty", "c_stm_quantity",
        ],
    },
    {
        "name": "CUSTOMER",
        "data_attr": "customers",
        "columns": [
            "NAME", "GROUP_NAME", "DESCRIPTION", "COMPANY_NAME",
            "ADDRESS1", "ADDRESS2", "ADDRESS3", "ADDRESS4", "ADDRESS5", "ADDRESS6",
            "FAX_NUM", "PHONE_NUM", "CONTACT", "EMAIL_ADDR",
        ],
        "field_map": [
            "name", "group_name", "description", "company_name",
            "address1", "address2", "address3", "address4", "address5", "address6",
            "fax_num", "phone_num", "contact", "email_addr",
        ],
    },
    {
        "name": "T_SITE",
        "data_attr": "t_sites",
        "columns": ["NAME", "DESCRIPTION", "GROUP_NAME", "PARENT_SITE"],
        "field_map": ["name", "description", "group_name", "parent_site"],
    },
    {
        "name": "T_PLANT",
        "data_attr": "t_plants",
        "columns": ["NAME", "DESCRIPTION", "GROUP_NAME", "SITE", "PERSONNEL_SMP_TYPE", "PERSONNEL_STAGE", "PERSONNEL_SPEC_TYPE"],
        "field_map": ["name", "description", "group_name", "site", "personnel_smp_type", "personnel_stage", "personnel_spec_type"],
    },
    {
        "name": "T_SUITE",
        "data_attr": "t_suites",
        "columns": ["NAME", "DESCRIPTION", "GROUP_NAME", "PLANT", "VISUAL_WORKFLOW", "CORRECTIVE_ACTION", "RESTRICT_COLLECTION"],
        "field_map": ["name", "description", "group_name", "plant", "visual_workflow", "corrective_action", "restrict_collection"],
    },
    {
        "name": "PROCESS_UNIT",
        "data_attr": "process_units",
        "columns": [
            "NAME", "DESCRIPTION", "GROUP_NAME", "PRODUCT_GRADE",
            "SAMPLE_TEMPLATE", "RUNNING", "PHASE", "DEFAULT_PHASE",
            "T_ALTERNATE_GRADE", "T_CORRECTIVE_ACTION", "T_SUITE",
        ],
        "field_map": [
            "name", "description", "group_name", "product_grade",
            "sample_template", "running", "phase", "default_phase",
            "t_alternate_grade", "t_corrective_action", "t_suite",
        ],
    },
    {
        "name": "PROC_SCHED_PARENT",
        "data_attr": "proc_sched_parents",
        "columns": [
            "NAME", "FIRST_DAY_OF_WEEK", "TREAT_HOLIDAYS_AS", "WORKSTATION",
            "RUNNING", "DESCRIPTION", "GROUP_NAME", "ACTIVE_FLAG", "VERSION", "T_SITE",
        ],
        "field_map": [
            "name", "first_day_of_week", "treat_holidays_as", "workstation",
            "running", "description", "group_name", "active_flag", "version", "t_site",
        ],
    },
    {
        "name": "PROCESS_SCHEDULE",
        "data_attr": "process_schedules",
        "columns": [
            "SCHEDULE_NUMBER", "LOGIN_OFFSET", "SCHD_COLLECT_TIME", "UNIT",
            "SAMPLING_POINT", "TEST_LIST", "ANALYSES",
            "MON_COLLECT", "TUE_COLLECT", "WED_COLLECT", "THU_COLLECT",
            "FRI_COLLECT", "SAT_COLLECT", "SUN_COLLECT",
            "WK_1_COLLECT", "WK_2_COLLECT", "WK_3_COLLECT", "WK_4_COLLECT",
            "DESCRIPTION", "SCHEDULE_NAME", "ORDER_NUMBER", "RUNNING", "PHASE",
            "SCHED_RULE", "WK_5_COLLECT", "DAY_COLLECT",
            "JAN_COLLECT", "FEB_COLLECT", "MAR_COLLECT", "APR_COLLECT",
            "MAY_COLLECT", "JUN_COLLECT", "JUL_COLLECT", "AUG_COLLECT",
            "SEP_COLLECT", "OCT_COLLECT", "NOV_COLLECT", "DEC_COLLECT",
            "DAYS_AHEAD", "R4_BASE_DATE", "R4_DAYS", "VERSION",
            "SPEC_TYPE", "STAGE", "T_COMPOSITE_GROUP", "T_CORRECTIVE_ACTION",
            "T_GROUP_ORDER", "T_MON_TYPE", "T_SAMPLE_TYPE", "T_TIME_WINDOW",
        ],
        "field_map": [
            "schedule_number", "login_offset", "schd_collect_time", "unit",
            "sampling_point", "test_list", "analyses",
            "mon_collect", "tue_collect", "wed_collect", "thu_collect",
            "fri_collect", "sat_collect", "sun_collect",
            "wk_1_collect", "wk_2_collect", "wk_3_collect", "wk_4_collect",
            "description", "schedule_name", "order_number", "running", "phase",
            "sched_rule", "wk_5_collect", "day_collect",
            "jan_collect", "feb_collect", "mar_collect", "apr_collect",
            "may_collect", "jun_collect", "jul_collect", "aug_collect",
            "sep_collect", "oct_collect", "nov_collect", "dec_collect",
            "days_ahead", "r4_base_date", "r4_days", "version",
            "spec_type", "stage", "t_composite_group", "t_corrective_action",
            "t_group_order", "t_mon_type", "t_sample_type", "t_time_window",
        ],
    },
    {
        "name": "LIST",
        "data_attr": "lists",
        "columns": ["LIST", "NAME", "VALUE", "ORDER_NUMBER"],
        "field_map": ["list_name", "name", "value", "order_number"],
    },
    {
        "name": "LIST_ENTRY",
        "data_attr": "list_entries",
        "columns": ["NAME", "GROUP_NAME", "DESCRIPTION"],
        "field_map": ["name", "group_name", "description"],
    },
    {
        "name": "VENDOR",
        "data_attr": "vendors",
        "columns": [
            "NAME", "DESCRIPTION", "CONTACT", "COMPANY_NAME",
            "ADDRESS1", "ADDRESS2", "ADDRESS3", "ADDRESS4", "ADDRESS5", "ADDRESS6",
            "FAX_NUM", "PHONE_NUM", "GROUP_NAME",
        ],
        "field_map": [
            "name", "description", "contact", "company_name",
            "address1", "address2", "address3", "address4", "address5", "address6",
            "fax_num", "phone_num", "group_name",
        ],
    },
    {
        "name": "SUPPLIER",
        "data_attr": "suppliers",
        "columns": [
            "NAME", "DESCRIPTION", "CONTACT", "COMPANY_NAME",
            "ADDRESS1", "ADDRESS2", "ADDRESS3", "ADDRESS4", "ADDRESS5", "ADDRESS6",
            "FAX_NUM", "PHONE_NUM", "GROUP_NAME",
        ],
        "field_map": [
            "name", "description", "contact", "company_name",
            "address1", "address2", "address3", "address4", "address5", "address6",
            "fax_num", "phone_num", "group_name",
        ],
    },
    {
        "name": "INSTRUMENTS",
        "data_attr": "instruments",
        "columns": [
            "NAME", "GROUP_NAME", "DESCRIPTION", "INST_GROUP", "VENDOR",
            "ON_LINE", "SERIAL_NO", "PM_DATE", "PM_INTV",
            "CALIB_DATE", "CALIB_INTV", "MODEL_NO",
        ],
        "field_map": [
            "name", "group_name", "description", "inst_group", "vendor",
            "on_line", "serial_no", "pm_date", "pm_intv",
            "calib_date", "calib_intv", "model_no",
        ],
    },
    {
        "name": "LIMS_USERS",
        "data_attr": "lims_users",
        "columns": [
            "USER_NAME", "FULL_NAME", "PHONE", "GROUP_NAME", "DESCRIPTION",
            "EMAIL_ADDR", "IS_ROLE", "USES_ROLES", "LANGUAGE_PREFIX", "T_SITE", "LOCATION_LAB",
        ],
        "field_map": [
            "user_name", "full_name", "phone", "group_name", "description",
            "email_addr", "is_role", "uses_roles", "language_prefix", "t_site", "location_lab",
        ],
    },
    {
        "name": "VERSIONS",
        "data_attr": "versions",
        "columns": ["TABLE_NAME", "VERSION", "DESCRIPTION"],
        "field_map": ["table_name", "version", "description"],
    },
]


class ExcelGenerator:
    """Generate a styled, multi-sheet LIMS Load Sheet Excel workbook."""

    def generate(
        self,
        result: ExtractionResult,
        output_path: str | Path,
        validation_issues: list[dict] | None = None,
    ) -> Path:
        output_path = Path(output_path)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        issue_index = self._build_issue_index(validation_issues or [])

        record_counts = []
        for spec in SHEET_SPECS:
            rows = getattr(result, spec["data_attr"], [])
            self._write_sheet(wb, spec, rows, issue_index)
            record_counts.append((spec["name"], len(rows)))

        # Summary sheet at position 0
        self._write_summary(wb, result, record_counts)

        wb.save(str(output_path))
        logger.info("Excel workbook saved to %s (%d sheets)", output_path, len(wb.sheetnames))
        return output_path

    def _write_sheet(
        self,
        wb: openpyxl.Workbook,
        spec: dict[str, Any],
        records: list[Any],
        issue_index: dict[str, set[tuple[int, str]]],
    ) -> None:
        ws = wb.create_sheet(title=spec["name"])

        # Header row
        for col_idx, col_name in enumerate(spec["columns"], start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _CELL_BORDER

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        # Data rows
        sheet_issues = issue_index.get(spec["name"], set())

        for row_idx, record in enumerate(records, start=2):
            is_low_conf = getattr(record, "confidence", 1.0) < 0.6

            for col_idx, field in enumerate(spec["field_map"], start=1):
                value = getattr(record, field, None)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = _NORMAL_FONT
                cell.border = _CELL_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)

                col_name = spec["columns"][col_idx - 1]
                if (row_idx - 2, col_name) in sheet_issues:
                    cell.fill = _ERROR_FILL
                elif is_low_conf:
                    cell.fill = _LOW_CONF_FILL

        self._auto_width(ws, spec["columns"])

    def _write_summary(
        self,
        wb: openpyxl.Workbook,
        result: ExtractionResult,
        record_counts: list[tuple[str, int]],
    ) -> None:
        ws = wb.create_sheet(title="Summary", index=0)

        # Header section
        header_rows = [
            ("LIMS Load Sheet", ""),
            ("Document Name", result.document_name),
            ("Document Type", result.document_type),
            ("Job ID", result.job_id),
            ("Overall Confidence", f"{result.overall_confidence:.1%}"),
            ("", ""),
            ("No.", "Table Name", "No. of Records"),
        ]

        for r_idx, row_data in enumerate(header_rows, start=1):
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = Font(bold=(c_idx == 1 or r_idx == 7), name="Calibri")

        # Table counts
        data_start = len(header_rows) + 1
        for i, (sheet_name, count) in enumerate(record_counts):
            row = data_start + i
            ws.cell(row=row, column=1, value=i + 1).font = _NORMAL_FONT
            ws.cell(row=row, column=2, value=sheet_name).font = _NORMAL_FONT
            ws.cell(row=row, column=3, value=count).font = _NORMAL_FONT

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15

    @staticmethod
    def _auto_width(ws: Any, columns: list[str]) -> None:
        for col_idx, header in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_width = len(header) + 4
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_width = max(max_width, min(len(str(cell.value)) + 2, 50))
            ws.column_dimensions[col_letter].width = max_width

    @staticmethod
    def _build_issue_index(issues: list[dict]) -> dict[str, set[tuple[int, str]]]:
        index: dict[str, set] = {}
        for issue in issues:
            sheet = issue.get("sheet", "")
            row = issue.get("row_index", -1)
            field = issue.get("field", "")
            if sheet:
                index.setdefault(sheet, set()).add((row, field))
        return index
