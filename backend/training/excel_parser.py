"""
excel_parser.py - Parse a completed LIMS Load Sheet Excel file into
a structured training example the LLM can read as few-shot context.

Each sheet is read and the first MAX_EXAMPLE_ROWS rows are kept
as concrete examples of correct mappings.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

MAX_EXAMPLE_ROWS = 5   # rows per sheet shown to the LLM (keep prompt short)

KNOWN_SHEETS = [
    "ANALYSIS_TYPES", "COMMON_NAME", "ANALYSIS", "COMPONENT", "UNITS",
    "PRODUCT", "T_PH_GRADE", "SAMPLING_POINT", "PRODUCT_GRADE",
    "PROD_GRADE_STAGE", "PRODUCT_SPEC", "T_PH_ITEM_CODE",
    "T_PH_ITEM_CODE_SPEC", "T_PH_ITEM_CODE_SUPP",
    "T_PH_SAMPLE_PLAN", "T_PH_SAMPLE_PLAN_EN",
    "CUSTOMER", "T_SITE", "T_PLANT", "T_SUITE", "PROCESS_UNIT",
    "PROC_SCHED_PARENT", "PROCESS_SCHEDULE", "LIST", "LIST_ENTRY",
    "VENDOR", "SUPPLIER", "INSTRUMENTS", "LIMS_USERS", "VERSIONS",
]


def parse_training_excel(file_path: str | Path) -> dict[str, Any]:
    """
    Parse a completed Load Sheet Excel file.

    Returns:
        {
          "sheets": {
              "Analysis": {
                  "columns": ["Name", "Group Name", ...],
                  "examples": [{"Name": "MOISTURE", ...}, ...]
              },
              ...
          },
          "sheet_count": 5,
          "total_rows": 42
        }
    """
    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    result: dict[str, Any] = {"sheets": {}, "sheet_count": 0, "total_rows": 0}

    for sheet_name in wb.sheetnames:
        if sheet_name == "Summary":
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # First row = headers
        headers = [str(c).strip() if c is not None else "" for c in rows[0]]
        if not any(headers):
            continue

        # Data rows (up to MAX_EXAMPLE_ROWS)
        examples = []
        for row in rows[1: MAX_EXAMPLE_ROWS + 1]:
            if not any(cell is not None for cell in row):
                continue
            record = {}
            for col_idx, header in enumerate(headers):
                if header:
                    val = row[col_idx] if col_idx < len(row) else None
                    record[header] = str(val).strip() if val is not None else ""
            if record:
                examples.append(record)

        if examples:
            result["sheets"][sheet_name] = {
                "columns": [h for h in headers if h],
                "examples": examples,
                "total_rows": len(rows) - 1,
            }
            result["total_rows"] += len(rows) - 1

    result["sheet_count"] = len(result["sheets"])
    wb.close()
    return result


def build_training_prompt_context(examples: list[dict[str, Any]]) -> str:
    """
    Format stored training examples into a prompt string injected
    before the extraction prompt so the LLM sees correct mappings.
    """
    if not examples:
        return ""

    lines = [
        "=== TRAINING EXAMPLES ===",
        "The following are real completed LIMS Load Sheet examples.",
        "Use these as a reference for how to map document content to the correct fields.",
        "",
    ]

    for ex in examples:
        lines.append(f"--- Example: {ex['name']} ---")
        if ex.get("description"):
            lines.append(f"Description: {ex['description']}")
        lines.append("")

        for sheet_name, sheet_data in ex.get("parsed_content", {}).get("sheets", {}).items():
            cols = sheet_data.get("columns", [])
            sample_rows = sheet_data.get("examples", [])
            if not sample_rows:
                continue
            lines.append(f"[{sheet_name}] columns: {', '.join(cols)}")
            for row in sample_rows[:3]:  # max 3 rows per sheet in prompt
                lines.append("  " + " | ".join(f"{k}: {v}" for k, v in row.items() if v))
            lines.append("")

    lines.append("=== END TRAINING EXAMPLES ===\n")
    return "\n".join(lines)
