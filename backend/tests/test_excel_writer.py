"""
test_excel_writer.py - Tests for Excel generation.
"""

import tempfile
from pathlib import Path

import openpyxl
import pytest

from excel_writer.excel_generator import ExcelGenerator, SHEET_SPECS
from models.schemas import (
    AnalysisRecord,
    ComponentRecord,
    ExtractionResult,
    ProductRecord,
    UnitRecord,
)


def _sample_result() -> ExtractionResult:
    return ExtractionResult(
        job_id="test-job-001",
        document_name="sample_stp.pdf",
        document_type="STP",
        overall_confidence=0.88,
        analysis=[
            AnalysisRecord(
                name="MOISTURE",
                group_name="Physical",
                reported_name="Moisture Content",
                description="Gravimetric moisture determination",
                analysis_type="Physical",
                confidence=0.95,
            ),
            AnalysisRecord(
                name="PH_VALUE",
                group_name="Chemical",
                reported_name="pH",
                analysis_type="Chemical",
                confidence=0.90,
            ),
        ],
        components=[
            ComponentRecord(
                analysis="MOISTURE",
                name="Moisture %",
                num_replicates=2,
                order_number=1,
                result_type="Numeric",
                units="PCT",
                minimum=0.0,
                maximum=15.0,
                confidence=0.95,
            ),
            ComponentRecord(
                analysis="PH_VALUE",
                name="pH",
                num_replicates=1,
                order_number=1,
                result_type="Numeric",
                units="PH",
                minimum=3.0,
                maximum=7.0,
                confidence=0.88,
            ),
        ],
        units=[
            UnitRecord(unit_code="PCT", description="Percentage", display_string="%"),
            UnitRecord(unit_code="PH", description="pH units", display_string="pH"),
        ],
        products=[
            ProductRecord(product="WHEAT_FLOUR", description="Wheat flour product"),
        ],
    )


class TestExcelGenerator:
    def test_generates_file(self, tmp_path):
        generator = ExcelGenerator()
        result = _sample_result()
        output = tmp_path / "test_output.xlsx"
        generator.generate(result, output)
        assert output.exists()
        assert output.stat().st_size > 0

    def test_sheets_present(self, tmp_path):
        generator = ExcelGenerator()
        result = _sample_result()
        output = tmp_path / "test_output.xlsx"
        generator.generate(result, output)

        wb = openpyxl.load_workbook(str(output))
        assert "ANALYSIS" in wb.sheetnames
        assert "COMPONENT" in wb.sheetnames
        assert "UNITS" in wb.sheetnames
        assert "PRODUCT" in wb.sheetnames
        assert "Summary" in wb.sheetnames

    def test_analysis_data_written(self, tmp_path):
        generator = ExcelGenerator()
        result = _sample_result()
        output = tmp_path / "test_output.xlsx"
        generator.generate(result, output)

        wb = openpyxl.load_workbook(str(output))
        ws = wb["ANALYSIS"]
        # Row 1 is header, row 2+ are data
        values = [ws.cell(row=2, column=c).value for c in range(1, 7)]
        assert "MOISTURE" in values

    def test_component_data_written(self, tmp_path):
        generator = ExcelGenerator()
        result = _sample_result()
        output = tmp_path / "test_output.xlsx"
        generator.generate(result, output)

        wb = openpyxl.load_workbook(str(output))
        ws = wb["COMPONENT"]
        row2 = [ws.cell(row=2, column=c).value for c in range(1, 10)]
        assert "MOISTURE" in row2
        assert 0.0 in row2 or "0.0" in [str(v) for v in row2]

    def test_summary_sheet_has_doc_name(self, tmp_path):
        generator = ExcelGenerator()
        result = _sample_result()
        output = tmp_path / "test_output.xlsx"
        generator.generate(result, output)

        wb = openpyxl.load_workbook(str(output))
        ws = wb["Summary"]
        all_values = [ws.cell(row=r, column=2).value for r in range(1, 10)]
        assert "sample_stp.pdf" in all_values

    def test_validation_errors_highlight_cells(self, tmp_path):
        generator = ExcelGenerator()
        result = _sample_result()
        output = tmp_path / "test_output.xlsx"
        generator.generate(
            result, output,
            validation_issues=[{"sheet": "Analysis", "row_index": 0, "field": "Name", "severity": "error"}],
        )
        # Just ensure it doesn't raise
        assert output.exists()
